import openpyxl

wb = openpyxl.load_workbook('tax_documents_sample.xlsx')
print('📊 Sheets in file:', wb.sheetnames)
print()

for sheet_name in wb.sheetnames[:2]:  # Show first 2 sheets
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:  # Show first 10 rows
            break
        print(row)