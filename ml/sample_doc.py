import random
from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl import Workbook
import requests
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
import os

# ---------- Helpers ----------
def money(x):
    d = Decimal(x).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return float(d)

def random_date(start_days_ago=365, end_days_ago=0):
    today = date.today()
    low = today - timedelta(days=start_days_ago)
    high = today - timedelta(days=end_days_ago)
    rand_days = random.randint(0, (high - low).days)
    return (low + timedelta(days=rand_days)).isoformat()

def generate_company():
    names = ["Acme Trading Pvt Ltd", "Nimbus Solutions", "Sunrise Exports", "BlueWave Imports"]
    return {
        "company_name": random.choice(names),
        "tax_id": f"GSTIN{random.randint(10000000,99999999)}",
    }

# ---------- Data Generators ----------
def generate_invoice():
    seller = generate_company()
    customer = generate_company()
    invoice_number = f"INV-{random.randint(1000,9999)}"
    inv_date = random_date()
    subtotal = 0
    items = []
    for i in range(random.randint(1,4)):
        qty = random.randint(1,10)
        rate = round(random.uniform(50, 500),2)
        line_total = qty * rate
        subtotal += line_total
        items.append((f"Item {i+1}", qty, rate, line_total))
    vat_rate = random.choice([0.05, 0.12, 0.18])
    vat_amount = round(subtotal * vat_rate,2)
    total = round(subtotal + vat_amount,2)
    return {
        "seller": seller,
        "customer": customer,
        "invoice_number": invoice_number,
        "invoice_date": inv_date,
        "items": items,
        "subtotal": subtotal,
        "vat_rate": vat_rate,
        "vat_amount": vat_amount,
        "total": total
    }

def generate_tax_return():
    company = generate_company()
    filing_date = random_date()
    turnover = round(random.uniform(10000,500000),2)
    vat_paid = round(turnover*0.12,2)
    refund_claim = round(max(0, vat_paid-random.uniform(0,vat_paid)),2)
    return {
        "company": company,
        "filing_date": filing_date,
        "turnover": turnover,
        "vat_paid": vat_paid,
        "refund_claim": refund_claim
    }

def generate_financial_statement():
    company = generate_company()
    revenue = round(random.uniform(50000,1000000),2)
    cogs = round(revenue*random.uniform(0.2,0.6),2)
    gross_profit = revenue-cogs
    opex = round(random.uniform(5000,gross_profit*0.5),2)
    net_profit = gross_profit-opex
    assets = round(random.uniform(50000,200000),2)
    liabilities = round(random.uniform(10000,assets*0.6),2)
    equity = assets-liabilities
    return {
        "company": company,
        "revenue": revenue,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "opex": opex,
        "net_profit": net_profit,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity
    }

def generate_bank_statement(n=5):
    company = generate_company()
    txns = []
    balance = random.uniform(1000,50000)
    for i in range(n):
        t_date = random_date()
        desc = random.choice(["Payment Received","Supplier Payment","Bank Charges","Salary"])
        amt = round(random.uniform(100,5000),2)
        t_type = random.choice(["CR","DR"])
        if t_type=="DR":
            balance -= amt
        else:
            balance += amt
        txns.append((t_date,desc,t_type,amt,round(balance,2)))
    return {"company": company, "transactions": txns}

def generate_receipt():
    receipt_id = f"REC-{random.randint(10000,99999)}"
    r_date = random_date()
    payer = generate_company()
    payee = generate_company()
    amount = round(random.uniform(100,10000),2)
    vat = round(amount*0.12,2)
    total = amount+vat
    return {
        "receipt_id": receipt_id,
        "date": r_date,
        "payer": payer,
        "payee": payee,
        "amount": amount,
        "vat": vat,
        "total": total
    }

def create_invoice_excel(inv, filename):
    """Create an Excel file for a single invoice"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Header information
    ws.append(["TAX INVOICE"])
    ws.append([])
    ws.append(["Seller:", inv['seller']['company_name']])
    ws.append(["GSTIN:", inv['seller']['tax_id']])
    ws.append([])
    ws.append(["Customer:", inv['customer']['company_name']])
    ws.append(["GSTIN:", inv['customer']['tax_id']])
    ws.append([])
    ws.append(["Invoice Number:", inv['invoice_number']])
    ws.append(["Date:", inv['invoice_date']])
    ws.append([])
    
    # Items table
    ws.append(["Item", "Qty", "Rate", "Total"])
    for item in inv['items']:
        ws.append([item[0], item[1], item[2], item[3]])
    
    ws.append([])
    ws.append(["Subtotal:", "", "", inv['subtotal']])
    ws.append([f"VAT ({inv['vat_rate']*100}%):", "", "", inv['vat_amount']])
    ws.append(["Total:", "", "", inv['total']])
    
    wb.save(filename)

# ---------- Generate Excel Files and Upload to Backend ----------
# Create sample_documents folder if it doesn't exist
sample_dir = "sample_documents"
if not os.path.exists(sample_dir):
    os.makedirs(sample_dir)
    print(f"📁 Created {sample_dir}/ folder")

excel_files = []

# Generate invoice Excel files and upload
for i in range(5):
    inv = generate_invoice()
    filename = f"sample_invoice_{i+1}.xlsx"
    filepath = os.path.join(sample_dir, filename)
    create_invoice_excel(inv, filepath)
    excel_files.append(filepath)

    # Upload to backend
    try:
        with open(filepath, 'rb') as f:
            files = {'documents': (filename, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post('http://localhost:3001/api/process-document', files=files)
            if response.status_code == 200:
                print(f"✅ {filename} uploaded and processed successfully.")
            else:
                print(f"❌ Failed to upload {filename}: {response.status_code}")
    except Exception as e:
        print(f"❌ Error uploading {filename}: {e}")

print(f"\n📄 Excel files saved in: {sample_dir}/")

# Create a comprehensive Excel file with all document types
wb = Workbook()

# Invoices
ws1 = wb.active
ws1.title = "Invoices"
ws1.append(["Invoice No","Date","Seller","Seller TaxID","Customer","Customer TaxID","Description","Qty","Rate","Line Total","Subtotal","VAT %","VAT Amt","Total"])
for _ in range(5):
    inv = generate_invoice()
    for item in inv["items"]:
        ws1.append([inv["invoice_number"],inv["invoice_date"],
                    inv["seller"]["company_name"],inv["seller"]["tax_id"],
                    inv["customer"]["company_name"],inv["customer"]["tax_id"],
                    item[0],item[1],item[2],item[3],
                    inv["subtotal"],inv["vat_rate"],inv["vat_amount"],inv["total"]])

# Tax Returns
ws2 = wb.create_sheet("Tax Returns")
ws2.append(["Company","Tax ID","Filing Date","Turnover","VAT Paid","Refund Claim"])
for _ in range(5):
    tr = generate_tax_return()
    ws2.append([tr["company"]["company_name"],tr["company"]["tax_id"],tr["filing_date"],tr["turnover"],tr["vat_paid"],tr["refund_claim"]])

# Financial Statements
ws3 = wb.create_sheet("Financial Statements")
ws3.append(["Company","Tax ID","Revenue","COGS","Gross Profit","Operating Expenses","Net Profit","Assets","Liabilities","Equity"])
for _ in range(5):
    fs = generate_financial_statement()
    ws3.append([fs["company"]["company_name"],fs["company"]["tax_id"],fs["revenue"],fs["cogs"],fs["gross_profit"],fs["opex"],fs["net_profit"],fs["assets"],fs["liabilities"],fs["equity"]])

# Bank Statements
ws4 = wb.create_sheet("Bank Statements")
ws4.append(["Company","Tax ID","Date","Description","Type","Amount","Running Balance"])
for _ in range(3):
    bs = generate_bank_statement(5)
    for t in bs["transactions"]:
        ws4.append([bs["company"]["company_name"],bs["company"]["tax_id"],t[0],t[1],t[2],t[3],t[4]])

# Receipts
ws5 = wb.create_sheet("Receipts")
ws5.append(["Receipt ID","Date","Payer","Payer TaxID","Payee","Payee TaxID","Amount","VAT","Total"])
for _ in range(5):
    r = generate_receipt()
    ws5.append([r["receipt_id"],r["date"],r["payer"]["company_name"],r["payer"]["tax_id"],r["payee"]["company_name"],r["payee"]["tax_id"],r["amount"],r["vat"],r["total"]])

wb.save("tax_documents_sample.xlsx")
print("✅ All sample Excel files generated and uploaded to backend for processing!")
print(f"📊 Individual invoice files: {len(excel_files)} files in {sample_dir}/")
print(f"📊 Comprehensive file: tax_documents_sample.xlsx")
